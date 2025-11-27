from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse
from .models import Proveedor
from django import forms
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ProveedorForm(forms.ModelForm):
    rut_nif = forms.CharField(
        max_length=12,
        required=True,
        widget=forms.TextInput(attrs={'class': 'form-input', 'placeholder': '12345678-5'})
    )
    email = forms.EmailField(
        required=True,
        widget=forms.EmailInput(attrs={'class': 'form-input', 'placeholder': 'contacto@proveedor.cl'})
    )
    email_secundario = forms.EmailField(
        required=False,
        widget=forms.EmailInput(attrs={'class': 'form-input', 'placeholder': 'alternativo@proveedor.cl'})
    )

    class Meta:
        model = Proveedor
        fields = ['rut_nif', 'nombre', 'contacto', 'direccion', 'pais', 'email', 'email_secundario']
        widgets = {
            'nombre': forms.TextInput(attrs={'class': 'form-input', 'placeholder': 'Nombre del proveedor'}),
            'contacto': forms.TextInput(attrs={'class': 'form-input', 'placeholder': 'Persona de contacto'}),
            'direccion': forms.Textarea(attrs={'class': 'form-input', 'placeholder': 'Dirección completa', 'rows': 3}),
            'pais': forms.TextInput(attrs={'class': 'form-input', 'placeholder': 'Chile'}),
        }

@login_required
def lista_proveedores(request):
    """Vista para listar todos los proveedores"""
    proveedores_qs = Proveedor.objects.all().order_by('-id_proveedor')

    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        proveedores_qs = proveedores_qs.filter(nombre__icontains=search) | proveedores_qs.filter(rut_nif__icontains=search)

    # Paginación con selector de tamaño
    per_page_param = request.GET.get('per_page')
    if per_page_param:
        per_page = int(per_page_param)
        request.session['proveedores_per_page'] = per_page
    else:
        per_page = request.session.get('proveedores_per_page', 10)
        if isinstance(per_page, str):
            per_page = int(per_page)

    paginator = Paginator(proveedores_qs, per_page)
    page = request.GET.get('page') or 1
    proveedores = paginator.get_page(page)

    context = {
        'proveedores': proveedores,
        'search': search,
        'per_page': per_page,
        'proveedores_count': Proveedor.objects.count(),
        'proveedores_activos': Proveedor.objects.count(),  # Sin campo de estado, usar total
        'productos_proveedor': 0,
        'ordenes_pendientes': 0,
    }
    return render(request, 'dashboard/proveedores.html', context)

@login_required
def form_proveedor(request):
    """Vista para mostrar el formulario de proveedores con la lista al lado"""
    proveedores = Proveedor.objects.all().order_by('-id_proveedor')
    
    context = {
        'form': ProveedorForm(),
        'proveedores': proveedores,
        'title': 'Formulario de Proveedor'
    }
    return render(request, 'dashboard/form_proveedor.html', context)

@login_required
def agregar_proveedor(request):
    """Vista para agregar un nuevo proveedor con diseño de 3 pasos"""
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            proveedor = form.save()
            
            # Si es una petición AJAX, responder con JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'proveedor': {
                        'id': proveedor.id_proveedor,
                        'nombre': proveedor.nombre,
                        'contacto': proveedor.contacto or '',
                        'direccion': proveedor.direccion or '',
                        'rut_nif': form.cleaned_data.get('rut_nif', ''),
                        'email': form.cleaned_data.get('email', ''),
                        'email_secundario': form.cleaned_data.get('email_secundario', ''),
                        'estado': 'ACTIVO'
                    }
                })
            
            messages.success(request, f'Proveedor "{proveedor.nombre}" creado exitosamente.')
            return redirect('proveedores:lista_proveedores')
        else:
            # Si hay errores y es AJAX, responder con JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
    else:
        form = ProveedorForm()
    
    # Obtener todos los proveedores para mostrar en la lista
    proveedores = Proveedor.objects.all().order_by('-id_proveedor')
    
    context = {
        'form': form,
        'proveedores': proveedores,
        'title': 'Agregar Proveedor',
        'action': 'agregar'
    }
    return render(request, 'dashboard/form_proveedor.html', context)

@login_required
def editar_proveedor(request, proveedor_id):
    """Vista para editar un proveedor existente"""
    proveedor = get_object_or_404(Proveedor, pk=proveedor_id)
    
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            proveedor = form.save()
            messages.success(request, f'Proveedor "{proveedor.nombre}" actualizado exitosamente.')
            return redirect('proveedores:lista_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    
    context = {
        'form': form,
        'proveedor': proveedor,
        'title': 'Editar Proveedor',
        'action': 'editar'
    }
    return render(request, 'dashboard/form_proveedor.html', context)

@login_required
def exportar_proveedores_excel(request):
    """Exporta proveedores (respetando filtros actuales) a Excel"""
    # Solo administradores pueden acceder
    user = request.user
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)

    # Filtros: actualmente soportamos 'search' por nombre/RUT
    search = request.GET.get('search', '')
    proveedores_qs = Proveedor.objects.all().order_by('-id_proveedor')
    if search:
        proveedores_qs = proveedores_qs.filter(nombre__icontains=search) | proveedores_qs.filter(rut_nif__icontains=search)

    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    # Estilos
    header_fill = PatternFill(start_color="4F81F7", end_color="4F81F7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados (alineados con la grilla visible)
    headers = ['ID', 'Nombre', 'Contacto', 'Dirección', 'País', 'RUT', 'Email', 'Email Secundario']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Llenar datos
    for row_num, prov in enumerate(proveedores_qs, 2):
        ws.cell(row=row_num, column=1).value = prov.id_proveedor
        ws.cell(row=row_num, column=2).value = prov.nombre
        ws.cell(row=row_num, column=3).value = prov.contacto or ''
        ws.cell(row=row_num, column=4).value = prov.direccion or ''
        ws.cell(row=row_num, column=5).value = prov.pais or ''
        ws.cell(row=row_num, column=6).value = prov.rut_nif
        ws.cell(row=row_num, column=7).value = prov.email or ''
        ws.cell(row=row_num, column=8).value = prov.email_secundario or ''
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_num).border = border
            ws.cell(row=row_num, column=col_num).alignment = Alignment(vertical='center')

    # Ajuste de columnas
    column_widths = [8, 30, 25, 40, 15, 15, 30, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'
    wb.save(response)
    return response

@login_required
def eliminar_proveedor(request, proveedor_id):
    """Vista para eliminar un proveedor"""
    proveedor = get_object_or_404(Proveedor, pk=proveedor_id)
    
    if request.method == 'POST':
        nombre = proveedor.nombre
        proveedor.delete()
        
        # Si es una petición AJAX, responder con JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Proveedor "{nombre}" eliminado exitosamente.'
            })
        
        messages.success(request, f'Proveedor "{nombre}" eliminado exitosamente.')
        return redirect('proveedores:lista_proveedores')
    
    return redirect('proveedores:lista_proveedores')
