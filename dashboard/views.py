from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.views.decorators.cache import never_cache
from django.views.decorators.debug import sensitive_post_parameters
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- API para eliminar producto ---
@login_required
@require_POST
def eliminar_producto(request, producto_id):
    user = request.user
    # Solo administradores pueden eliminar
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    try:
        producto = get_object_or_404(Producto, id_producto=producto_id)
        nombre = producto.nombre
        producto.delete()
        # Auditoría
        Auditoria.objects.create(
            usuario=user,
            accion='BORRAR',
            entidad='Producto',
            detalle=f'ID: {producto_id}, Nombre: {nombre}'
        )
        return JsonResponse({'success': True, 'message': f'Producto "{nombre}" eliminado correctamente'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# --- API para eliminar inventario ---
@login_required
@require_POST
def eliminar_inventario(request, inventario_id):
    user = request.user
    rol_nombre = user.id_rol.nombre if hasattr(user, 'id_rol') and user.id_rol else None
    # Solo administradores y bodegueros pueden eliminar
    if not (user.is_superuser or rol_nombre in ['Administrador', 'Bodeguero']):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    try:
        inventario = get_object_or_404(Inventario, id_inventario=inventario_id)
        producto_nombre = inventario.id_producto.nombre
        ubicacion = inventario.ubicacion
        inventario.delete()
        # Auditoría
        Auditoria.objects.create(
            usuario=user,
            accion='BORRAR',
            entidad='Inventario',
            detalle=f'ID: {inventario_id}, Producto: {producto_nombre}, Ubicación: {ubicacion}'
        )
        return JsonResponse({'success': True, 'message': f'Inventario de "{producto_nombre}" eliminado correctamente'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
from datetime import datetime
from django.contrib.auth.hashers import make_password
from django.conf import settings
from productos.models import Producto
from inventarios.models import Inventario
from usuarios.models import Usuario, PasswordResetToken
from .forms import ProductoForm, InventarioForm
from .models import Auditoria

@never_cache
@sensitive_post_parameters('password')
def login_view(request):
    """Vista de login personalizada con protección anti-fuerza bruta"""
    # Si el usuario ya está autenticado, redirigir al dashboard
    if request.user.is_authenticated:
        return redirect('dashboard:home')
    
    if request.method == 'POST':
        identificador = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        
        # F-LOGIN-05: Validar campos obligatorios antes de procesar
        if not identificador or not password:
            context = {
                'error_type': 'empty_fields',
                'username': identificador
            }
            return render(request, 'dashboard/new_login.html', context)
        # Permitir login por username o correo
        try:
            if '@' in identificador:
                user_obj = Usuario.objects.get(correo=identificador)
            else:
                user_obj = Usuario.objects.get(username=identificador)
            username_real = user_obj.username
            
            # Verificar si la cuenta está bloqueada
            if user_obj.is_account_locked():
                tiempo_restante = int((user_obj.locked_until - timezone.now()).total_seconds() / 60)
                context = {
                    'error_type': 'locked',
                    'locked_minutes': tiempo_restante,
                    'username': identificador
                }
                return render(request, 'dashboard/new_login.html', context)
            
            # Intentar autenticar
            user = authenticate(request, username=username_real, password=password)
            
            if user is not None:
                # Login exitoso - resetear intentos fallidos
                user_obj.reset_failed_attempts()
                login(request, user)
                # Si el usuario debe cambiar la clave, redirigir al flujo de cambio
                if hasattr(user_obj, 'debe_cambiar_clave') and user_obj.debe_cambiar_clave:
                    return redirect('dashboard:reset_password')
                return redirect('dashboard:home')
            else:
                # Contraseña incorrecta - incrementar intentos fallidos
                user_obj.increment_failed_attempts()
                intentos_restantes = 5 - user_obj.failed_login_attempts
                
                if user_obj.is_account_locked():
                    # Cuenta bloqueada
                    context = {
                        'error_type': 'locked',
                        'locked_minutes': 30,
                        'username': identificador
                    }
                else:
                    # Advertencia de intentos restantes
                    context = {
                        'error_type': 'invalid',
                        'attempts_remaining': intentos_restantes,
                        'username': identificador
                    }
                
                return render(request, 'dashboard/new_login.html', context)
        
        except Usuario.DoesNotExist:
            # Usuario/Correo no existe - no dar pistas de seguridad
            context = {
                'error_type': 'invalid',
                'username': identificador
            }
            return render(request, 'dashboard/new_login.html', context)
    
    return render(request, 'dashboard/new_login.html')

def logout_view(request):
    """Vista de logout con protección anti-caché (F-LOGIN-06)"""
    # Destruir sesión completamente
    if request.user.is_authenticated:
        logout(request)
    
    # Limpiar cookies de sesión
    request.session.flush()
    
    # Crear respuesta con headers anti-caché y JavaScript para limpiar historial
    response = redirect('login')
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    # Agregar header para prevenir navegación hacia atrás
    response['Clear-Site-Data'] = '"cache", "cookies", "storage"'
    
    return response

@login_required
@never_cache
def home(request):
    """Dashboard principal"""
    user = request.user
    now = timezone.now()
    
    # Datos para el contexto
    context = {
        'user': user,
        'productos_count': Producto.objects.count(),
        'inventarios_count': Inventario.objects.count(),
        'today': now.date(),
        'now': now,
    }
    
    # Datos ficticios para proveedores y ventas (solo para administradores)
    if user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador'):
        context.update({
            'proveedores_count': 12,  # Ficticio
            'ventas_count': 156,      # Ficticio
        })
    
    return render(request, 'dashboard/home.html', context)

@login_required
@never_cache
def productos_view(request):
    """Vista para listar todos los productos con búsqueda, paginación y ordenamiento"""
    user = request.user
    
    # Verificar permisos según rol
    rol_nombre = user.id_rol.nombre if hasattr(user, 'id_rol') and user.id_rol else None
    es_vendedor = rol_nombre == 'Vendedor'
    es_bodeguero = rol_nombre == 'Bodeguero'
    puede_crear_editar = user.is_superuser or rol_nombre in ['Administrador', 'Bodeguero']
    puede_eliminar = user.is_superuser or rol_nombre == 'Administrador'
    
    productos = Producto.objects.all()
    
    # Búsqueda por múltiples campos
    search = request.GET.get('search', '')
    if search:
        productos = productos.filter(
            nombre__icontains=search
        ) | productos.filter(
            descripcion__icontains=search
        ) | productos.filter(
            precio_referencia__icontains=search
        )
    
    # Ordenamiento
    order_by = request.GET.get('order_by', 'id_producto')
    order_direction = request.GET.get('order_direction', 'asc')
    
    # Construir el campo de ordenamiento
    if order_direction == 'desc':
        order_field = f'-{order_by}' if not order_by.startswith('-') else order_by
    else:
        order_field = order_by.replace('-', '')
    
    productos = productos.order_by(order_field)
    
    # Paginación - obtener de sesión o de parámetro GET
    per_page_param = request.GET.get('per_page')
    if per_page_param:
        per_page = int(per_page_param)
        request.session['productos_per_page'] = per_page
    else:
        per_page = request.session.get('productos_per_page', 10)
        # Asegurar que sea entero
        if isinstance(per_page, str):
            per_page = int(per_page)
    
    paginator = Paginator(productos, per_page)
    page = request.GET.get('page', 1)
    productos_paginados = paginator.get_page(page)
    
    context = {
        'productos': productos_paginados,
        'search': search,
        'order_by': order_by.replace('-', ''),
        'order_direction': order_direction,
        'per_page': per_page,
        'total_productos': Producto.objects.count(),
        'productos_activos': Producto.objects.count(),
        'es_vendedor': es_vendedor,
        'es_bodeguero': es_bodeguero,
        'puede_crear_editar': puede_crear_editar,
        'puede_eliminar': puede_eliminar,
    }
    return render(request, 'dashboard/productos.html', context)

@login_required
@never_cache
def inventarios_view(request):
    """Vista de inventarios con paginación y selector de registros por página"""
    user = request.user
    # Verificar permisos según rol
    rol_nombre = user.id_rol.nombre if hasattr(user, 'id_rol') and user.id_rol else None
    # Bloquear acceso al rol CONSULTA
    if rol_nombre == 'Consulta':
        raise PermissionDenied("El rol Consulta no tiene permisos para acceder al inventario")
    es_vendedor = rol_nombre == 'Vendedor'
    es_bodeguero = rol_nombre == 'Bodeguero'
    puede_editar = user.is_superuser or rol_nombre in ['Administrador', 'Bodeguero']
    inventarios_qs = Inventario.objects.select_related('id_producto').all()
    now = timezone.now()
    # Mock data para proveedores
    class MockProveedor:
        def __init__(self, id_proveedor, nombre):
            self.id_proveedor = id_proveedor
            self.nombre = nombre
    proveedores = [
        MockProveedor(1, 'Distribuidora Nacional'),
        MockProveedor(2, 'Dulces Premium SAC'),
        MockProveedor(3, 'Confitería del Norte'),
    ]
    # Paginación - obtener de sesión o de parámetro GET
    per_page_param = request.GET.get('per_page')
    if per_page_param:
        per_page = int(per_page_param)
        request.session['inventarios_per_page'] = per_page
    else:
        per_page = request.session.get('inventarios_per_page', 10)
        if isinstance(per_page, str):
            per_page = int(per_page)
    paginator = Paginator(inventarios_qs, per_page)
    page = request.GET.get('page', 1)
    inventarios = paginator.get_page(page)
    context = {
        'inventarios': inventarios,
        'proveedores': proveedores,
        'total_productos': inventarios_qs.count(),
        'stock_alto': 0,  # Calcular basado en lógica de stock
        'stock_medio': 0,
        'stock_bajo': 0,
        'today': now.date(),
        'user': request.user,
        'es_vendedor': es_vendedor,
        'es_bodeguero': es_bodeguero,
        'puede_editar': puede_editar,
        'per_page': per_page,
    }
    return render(request, 'dashboard/inventarios.html', context)

@login_required
@never_cache
def proveedores_view(request):
    """Vista de proveedores con datos reales y paginación"""
    user = request.user
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        raise PermissionDenied("No tienes permisos para acceder a esta sección")

    # Datos reales de proveedores
    from proveedores.models import Proveedor
    proveedores_qs = Proveedor.objects.all().order_by('-id_proveedor')

    # Búsqueda (nombre o RUT)
    search = request.GET.get('search', '')
    if search:
        proveedores_qs = proveedores_qs.filter(nombre__icontains=search) | proveedores_qs.filter(rut_nif__icontains=search)

    # Paginación y selector de registros por página
    per_page_param = request.GET.get('per_page')
    if per_page_param:
        per_page = int(per_page_param)
        request.session['proveedores_per_page'] = per_page
    else:
        per_page = request.session.get('proveedores_per_page', 10)
        if isinstance(per_page, str):
            per_page = int(per_page)

    paginator = Paginator(proveedores_qs, per_page)
    page = request.GET.get('page', 1)
    proveedores = paginator.get_page(page)

    context = {
        'proveedores': proveedores,
        'search': search,
        'per_page': per_page,
        'proveedores_count': Proveedor.objects.count(),
        'proveedores_activos': Proveedor.objects.count(),  # sin campo de estado, usamos total
        'productos_proveedor': 0,
        'ordenes_pendientes': 0,
        'user': request.user,
    }
    return render(request, 'dashboard/proveedores.html', context)

@login_required
@never_cache
def ventas_view(request):
    """Vista ficticia de ventas"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        raise PermissionDenied("No tienes permisos para acceder a esta sección")
    
    # Datos ficticios
    ventas = [
        {'id': 1, 'fecha': '2025-01-15', 'cliente': 'Cliente A', 'total': 50000, 'estado': 'Completada'},
        {'id': 2, 'fecha': '2025-01-14', 'cliente': 'Cliente B', 'total': 75000, 'estado': 'Pendiente'},
        {'id': 3, 'fecha': '2025-01-13', 'cliente': 'Cliente C', 'total': 30000, 'estado': 'Completada'},
    ]
    
    context = {
        'ventas': ventas,
        'user': request.user,
    }
    return render(request, 'dashboard/ventas.html', context)

@login_required
def agregar_producto(request):
    """Vista para agregar un nuevo producto"""
    user = request.user
    
    # Verificar permisos
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre in ['Administrador', 'Bodeguero'])):
        messages.error(request, 'No tienes permisos para agregar productos')
        return redirect('dashboard:productos')
    
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save()
            # Auditoría
            Auditoria.objects.create(
                usuario=user,
                accion='CREAR',
                entidad='Producto',
                detalle=f'ID: {producto.id_producto}, Nombre: {producto.nombre}'
            )
            messages.success(request, f'Producto "{producto.nombre}" agregado exitosamente')
            return redirect('dashboard:productos')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario')
    else:
        form = ProductoForm()
    
    context = {
        'form': form,
        'user': user,
        'titulo': 'Agregar Nuevo Producto'
    }
    return render(request, 'dashboard/form_producto.html', context)

@login_required
def editar_producto(request, producto_id):
    """Vista para editar un producto existente"""
    user = request.user
    producto = get_object_or_404(Producto, id_producto=producto_id)
    
    # Solo administradores pueden editar
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        messages.error(request, 'No tienes permisos para editar productos')
        return redirect('dashboard:productos')
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            producto = form.save()
            # Auditoría
            Auditoria.objects.create(
                usuario=user,
                accion='EDITAR',
                entidad='Producto',
                detalle=f'ID: {producto.id_producto}, Nombre: {producto.nombre}'
            )
            messages.success(request, 'Producto actualizado exitosamente')
            return redirect('dashboard:productos')
    else:
        form = ProductoForm(instance=producto)
    
    context = {
        'form': form,
        'user': user,
        'titulo': f'Editar Producto: {producto.nombre}',
        'producto': producto
    }
    return render(request, 'dashboard/form_producto.html', context)

@login_required
def agregar_inventario(request):
    """Vista para agregar un nuevo inventario"""
    user = request.user
    rol_nombre = user.id_rol.nombre if hasattr(user, 'id_rol') and user.id_rol else None
    
    # Bloquear acceso a roles sin permisos de escritura (CONSULTA, VENDEDOR)
    if rol_nombre in ['Consulta', 'Vendedor']:
        raise PermissionDenied(f"El rol {rol_nombre} no tiene permisos para crear inventarios")
    
    # Solo administradores y bodegueros pueden agregar inventarios
    if not (user.is_superuser or rol_nombre in ['Administrador', 'Bodeguero']):
        raise PermissionDenied('No tienes permisos para agregar inventarios')
    
    if request.method == 'POST':
        form = InventarioForm(request.POST)
        if form.is_valid():
            inventario = form.save()
            # Auditoría
            Auditoria.objects.create(
                usuario=user,
                accion='CREAR',
                entidad='Inventario',
                detalle=f'ID: {inventario.id_inventario}, Producto: {inventario.id_producto.nombre}, Ubicación: {inventario.ubicacion}'
            )
            messages.success(request, 'Inventario agregado exitosamente')
            return redirect('dashboard:inventarios')
    else:
        form = InventarioForm()
    
    context = {
        'form': form,
        'user': user,
        'titulo': 'Agregar Nuevo Inventario'
    }
    return render(request, 'dashboard/form_inventario.html', context)

@login_required
def editar_inventario(request, inventario_id):
    """Vista para editar un inventario existente"""
    user = request.user
    inventario = get_object_or_404(Inventario, id_inventario=inventario_id)
    rol_nombre = user.id_rol.nombre if hasattr(user, 'id_rol') and user.id_rol else None
    
    # Bloquear acceso a roles sin permisos de escritura (CONSULTA, VENDEDOR)
    if rol_nombre in ['Consulta', 'Vendedor']:
        raise PermissionDenied(f"El rol {rol_nombre} no tiene permisos para editar inventarios")
    
    # Solo administradores y bodegueros pueden editar
    if not (user.is_superuser or rol_nombre in ['Administrador', 'Bodeguero']):
        raise PermissionDenied('No tienes permisos para editar inventarios')
    
    if request.method == 'POST':
        form = InventarioForm(request.POST, instance=inventario)
        if form.is_valid():
            inventario = form.save()
            # Auditoría
            Auditoria.objects.create(
                usuario=user,
                accion='EDITAR',
                entidad='Inventario',
                detalle=f'ID: {inventario.id_inventario}, Producto: {inventario.id_producto.nombre}, Ubicación: {inventario.ubicacion}'
            )
            messages.success(request, 'Inventario actualizado exitosamente')
            return redirect('dashboard:inventarios')
    else:
        form = InventarioForm(instance=inventario)
    
    context = {
        'form': form,
        'user': user,
        'titulo': f'Editar Inventario: {inventario.id_producto.nombre}',
        'inventario': inventario
    }
    return render(request, 'dashboard/form_inventario.html', context)

def forgot_password_view(request):
    """Vista para recuperación de contraseña"""
    if request.method == 'POST':
        email = request.POST.get('email')
        
        try:
            usuario = Usuario.objects.get(correo=email)
            
            # Crear token de recuperación
            token = PasswordResetToken.objects.create(usuario=usuario)
            
            # Calcular minutos de expiración
            expiration_minutes = int((token.expires_at - token.created_at).total_seconds() / 60)
            
            # Construir URL de recuperación
            reset_url = request.build_absolute_uri(
                f'/reset-password/?token={token.token}'
            )
            
            # Renderizar template de email
            html_message = render_to_string('dashboard/password_reset_email.html', {
                'usuario': usuario,
                'reset_url': reset_url,
                'token': token,
                'expiration_minutes': expiration_minutes
            })
            plain_message = strip_tags(html_message)
            
            # Enviar email
            try:
                send_mail(
                    subject='Recuperación de Contraseña - Dulcería Lilis',
                    message=plain_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    html_message=html_message,
                    fail_silently=False,
                )
                
                # Log exitoso (opcional)
                print(f"✅ Email de recuperación enviado a: {email}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': 'Se ha enviado un email con las instrucciones para recuperar tu contraseña'
                    })
                
                messages.success(request, 'Se ha enviado un email con las instrucciones para recuperar tu contraseña')
                return redirect('dashboard:login')
                
            except Exception as email_error:
                # Log del error
                import traceback
                print(f"❌ Error enviando email a {email}:")
                print(traceback.format_exc())
                
                # Marcar el token como inválido si el email falló
                token.delete()
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Error al enviar el email. Verifica la configuración del servidor de correo.'
                    })
                
                messages.error(request, 'Error al enviar el email. Por favor contacta al administrador del sistema.')
                return render(request, 'dashboard/forgot_password.html')
            
        except Usuario.DoesNotExist:
            # F-REC-02: Mostrar mensaje genérico sin indicar si el correo existe o no
            print(f"⚠️ Intento de recuperación para email no registrado: {email}")
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Si el email existe en nuestro sistema, recibirás un correo con instrucciones para recuperar tu contraseña. Por favor revisa tu bandeja de entrada y spam.'
                })
            
            messages.info(request, 'Si el email existe en nuestro sistema, recibirás un correo con instrucciones.')
            return redirect('dashboard:login')
            
        except Exception as e:
            import traceback
            print(f"❌ Error general en forgot_password_view:")
            print(traceback.format_exc())
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Error al procesar la solicitud. Por favor intenta nuevamente.'
                })
            
            messages.error(request, 'Error al procesar la solicitud. Por favor intenta nuevamente.')
    
    return render(request, 'dashboard/forgot_password.html')

@sensitive_post_parameters('password', 'password_confirm')
def reset_password_view(request):
    """Vista para resetear contraseña con token"""
    token_str = request.GET.get('token') or request.POST.get('token')
    
    # Flujo de primer ingreso: usuario autenticado con debe_cambiar_clave=True sin token
    if not token_str and request.user.is_authenticated:
        usuario = request.user
        if hasattr(usuario, 'debe_cambiar_clave') and usuario.debe_cambiar_clave:
            if request.method == 'POST':
                password = (request.POST.get('password') or '').strip()
                password_confirm = (request.POST.get('password_confirm') or '').strip()
                # Validaciones básicas de contraseña
                if password != password_confirm:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': 'Las contraseñas no coinciden'})
                    messages.error(request, 'Las contraseñas no coinciden')
                elif len(password) < 8:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': 'La contraseña debe tener al menos 8 caracteres'})
                    messages.error(request, 'La contraseña debe tener al menos 8 caracteres')
                elif not any(c.isupper() for c in password):
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': 'La contraseña debe contener al menos una letra mayúscula'})
                    messages.error(request, 'La contraseña debe contener al menos una letra mayúscula')
                elif not any(c.islower() for c in password):
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': 'La contraseña debe contener al menos una letra minúscula'})
                    messages.error(request, 'La contraseña debe contener al menos una letra minúscula')
                elif not any(c.isdigit() for c in password):
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': 'La contraseña debe contener al menos un número'})
                    messages.error(request, 'La contraseña debe contener al menos un número')
                elif not any(c in '!@#$%^&*()-_=+[]{};:,./?' for c in password):
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': 'La contraseña debe contener al menos un carácter especial'})
                    messages.error(request, 'La contraseña debe contener al menos un carácter especial')
                else:
                    try:
                        usuario.set_password(password)
                        usuario.debe_cambiar_clave = False
                        usuario.save(update_fields=['password', 'debe_cambiar_clave'])
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': True, 'message': 'Contraseña cambiada exitosamente'})
                        messages.success(request, 'Contraseña actualizada. ¡Bienvenido!')
                        return redirect('dashboard:home')
                    except Exception as e:
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'message': 'Error al cambiar la contraseña'}, status=500)
                        messages.error(request, 'Error al cambiar la contraseña')
            # Mostrar formulario de cambio de clave sin token
            context = {
                'token': None,
                'usuario': usuario,
                'primer_ingreso': True,
            }
            return render(request, 'dashboard/reset_password.html', context)
        # Si no requiere cambio, envía a home
        return redirect('dashboard:home')
    
    if not token_str:
        messages.error(request, 'Token de recuperación no válido')
        return redirect('dashboard:login')
    
    try:
        token = PasswordResetToken.objects.get(token=token_str)
        
        if not token.is_valid():
            messages.error(request, 'El token ha expirado o ya fue usado. Solicita uno nuevo.')
            return redirect('dashboard:forgot_password')
        
        if request.method == 'POST':
            password = (request.POST.get('password') or '').strip()
            password_confirm = (request.POST.get('password_confirm') or '').strip()
            
            # F-REC-05: Validación de contraseñas
            if password != password_confirm:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Las contraseñas no coinciden'
                    })
                messages.error(request, 'Las contraseñas no coinciden')
            elif len(password) < 8:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'La contraseña debe tener al menos 8 caracteres'
                    })
                messages.error(request, 'La contraseña debe tener al menos 8 caracteres')
            elif not any(c.isupper() for c in password):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'La contraseña debe contener al menos una letra mayúscula'
                    })
                messages.error(request, 'La contraseña debe contener al menos una letra mayúscula')
            elif not any(c.islower() for c in password):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'La contraseña debe contener al menos una letra minúscula'
                    })
                messages.error(request, 'La contraseña debe contener al menos una letra minúscula')
            elif not any(c.isdigit() for c in password):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'La contraseña debe contener al menos un número'
                    })
                messages.error(request, 'La contraseña debe contener al menos un número')
            elif not any(c in '!@#$%^&*()-_=+[]{};:,./?' for c in password):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'La contraseña debe contener al menos un carácter especial'
                    })
                messages.error(request, 'La contraseña debe contener al menos un carácter especial')
            else:
                # F-REC-03: Cambiar contraseña y marcar token como usado (usar set_password para coherencia)
                usuario = token.usuario
                usuario.set_password(password)
                usuario.save(update_fields=['password'])
                
                # Marcar token como usado
                token.is_used = True
                token.save()
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': 'Contraseña cambiada exitosamente'
                    })
                
                messages.success(request, 'Contraseña cambiada exitosamente. Ahora puedes iniciar sesión.')
                return redirect('dashboard:login')
        
        context = {
            'token': token_str,
            'usuario': token.usuario
        }
        return render(request, 'dashboard/reset_password.html', context)
        
    except PasswordResetToken.DoesNotExist:
        messages.error(request, 'Token de recuperación no válido')
        return redirect('dashboard:login')


@login_required
@never_cache
def usuarios_view(request):
    """Vista de gestión de usuarios con paginación y selector de registros por página"""
    user = request.user
    # Solo administradores pueden gestionar usuarios
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        raise PermissionDenied("No tienes permisos para gestionar usuarios")
    # Usar el modelo de Usuario personalizado
    from usuarios.models import Usuario
    from roles.models import Rol
    usuarios_qs = Usuario.objects.select_related('id_rol').all()
    roles = Rol.objects.all()
    
    # Paginación - obtener de sesión o de parámetro GET
    per_page_param = request.GET.get('per_page')
    if per_page_param:
        per_page = int(per_page_param)
        request.session['usuarios_per_page'] = per_page
    else:
        per_page = request.session.get('usuarios_per_page', 10)
        if isinstance(per_page, str):
            per_page = int(per_page)
    paginator = Paginator(usuarios_qs, per_page)
    page = request.GET.get('page', 1)
    usuarios = paginator.get_page(page)
    context = {
        'usuarios': usuarios,
        'roles': roles,
        'usuarios_count': Usuario.objects.count(),
        'usuarios_activos': Usuario.objects.filter(is_active=True).count(),
        'usuarios_inactivos': Usuario.objects.filter(is_active=False).count(),
        'nuevos_usuarios': 0,  # Mock data
        'user': request.user,
        'per_page': per_page,
    }
    return render(request, 'dashboard/usuarios.html', context)

@login_required
def obtener_usuario(request, usuario_id):
    """API para obtener datos de un usuario en formato JSON"""
    import traceback
    
    # Verificar autenticación
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'No autenticado'}, status=401)
    
    user = request.user
    
    # Solo administradores pueden acceder
    try:
        if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
            return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error verificando permisos: {str(e)}'}, status=500)
    
    try:
        from usuarios.models import Usuario
        usuario = Usuario.objects.select_related('id_rol').get(id_usuario=usuario_id)
        
        data = {
            'success': True,
            'usuario': {
                'id': usuario.id_usuario,
                'usuario': usuario.username,
                'email': usuario.email,
                'nombre': usuario.nombre,
                'telefono': usuario.telefono or '',
                'id_rol': usuario.id_rol.id_rol if usuario.id_rol else '',
                'is_active': usuario.is_active,
                'cambiar_password': False  # Por defecto no forzar cambio
            }
        }
        return JsonResponse(data)
    except Usuario.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"Error en obtener_usuario: {error_trace}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@login_required
def guardar_usuario(request):
    """API para crear o actualizar un usuario"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
    
    try:
        from usuarios.models import Usuario
        from roles.models import Rol
        
        usuario_id = request.POST.get('user_id')
        usuario_username = request.POST.get('usuario')
        email = request.POST.get('email')
        password = request.POST.get('password')
        nombre = request.POST.get('nombre')
        telefono = request.POST.get('telefono', '')
        id_rol = request.POST.get('id_rol')
        activo_str = request.POST.get('activo')
        
        # Convertir el valor del select a booleano
        if activo_str == 'true':
            activo = True
        elif activo_str == 'false':
            activo = False
        else:
            # Si no se seleccionó ningún estado, retornar error
            return JsonResponse({
                'success': False,
                'errors': {'activo': ['Debes seleccionar un estado (Activo o Inactivo)']}
            })
        
        # Validaciones básicas
        if not all([usuario_username, email, nombre, id_rol]):
            return JsonResponse({
                'success': False, 
                'errors': {'general': ['Todos los campos requeridos deben ser completados']}
            })
        
        # Obtener el rol
        try:
            rol = Rol.objects.get(pk=id_rol)
        except Rol.DoesNotExist:
            return JsonResponse({
                'success': False,
                'errors': {'id_rol': ['El rol seleccionado no existe']}
            })
        
        # Modo edición o creación
        if usuario_id:
            # Editar usuario existente
            usuario = get_object_or_404(Usuario, id_usuario=usuario_id)
            
            # Verificar que el username no esté siendo usado por otro usuario
            if Usuario.objects.filter(username=usuario_username).exclude(id_usuario=usuario_id).exists():
                return JsonResponse({
                    'success': False,
                    'errors': {'usuario': ['Este nombre de usuario ya está en uso']}
                })
            
            # Actualizar datos
            usuario.username = usuario_username
            usuario.email = email
            usuario.correo = email
            usuario.nombre = nombre
            usuario.telefono = telefono
            usuario.id_rol = rol
            usuario.is_active = activo
            
            # Reset de contraseña por administrador: generar temporal si se solicitó
            if password:
                usuario.set_password(password)
                usuario.debe_cambiar_clave = True
                usuario.save(update_fields=['password', 'debe_cambiar_clave'])
                try:
                    login_url = request.build_absolute_uri(reverse('dashboard:login'))
                    html_message = (
                        f"""
                        <div style='font-family:Segoe UI, Arial, sans-serif; color:#111; line-height:1.5;'>
                          <h2 style='margin:0 0 12px; color:#dc2626;'>Dulcería Lilis</h2>
                          <p>Hola <strong>{usuario.nombre}</strong>,</p>
                          <p>Se ha generado una clave temporal para tu cuenta.</p>
                          <div style='background:#f9fafb; border:1px solid #e5e7eb; border-radius:8px; padding:12px; margin:14px 0;'>
                            <p style='margin:0;'><strong>Usuario:</strong> {usuario.username}</p>
                            <p style='margin:6px 0 0;'><strong>Clave temporal:</strong> <code style='font-size:14px;'>{password}</code></p>
                          </div>
                          <p>Puedes ingresar desde este enlace:</p>
                          <p><a href='{login_url}' style='background:#dc2626; color:#fff; padding:10px 14px; border-radius:6px; text-decoration:none;'>Ir al acceso</a></p>
                          <p style='color:#6b7280; font-size:13px;'>Al ingresar se te pedirá cambiar la contraseña.</p>
                        </div>
                        """
                    )
                    send_mail(
                        subject='Dulcería Lilis - Clave temporal de acceso',
                        message=(
                            f'Hola {usuario.nombre},\n'
                            f'Se ha generado una clave temporal para tu cuenta.\n'
                            f'Usuario: {usuario.username}\n'
                            f'Clave temporal: {password}\n\n'
                            f'URL de acceso: {login_url}\n'
                            'Al ingresar se te pedirá cambiar la contraseña.'
                        ),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[usuario.email],
                        fail_silently=True,
                        html_message=html_message,
                    )
                except Exception:
                    pass
            
            usuario.save()
            # Auditoría
            Auditoria.objects.create(
                usuario=user,
                accion='EDITAR',
                entidad='Usuario',
                detalle=f'ID: {usuario.id_usuario}, Nombre: {usuario.nombre}'
            )
            action = 'actualizado'
        else:
            # Crear nuevo usuario
            # Verificar que el username no exista
            if Usuario.objects.filter(username=usuario_username).exists():
                return JsonResponse({
                    'success': False,
                    'errors': {'usuario': ['Este nombre de usuario ya está en uso']}
                })
            
            # Generar clave temporal robusta si no se proporcionó
            import random, string
            def generar_clave_temporal(longitud=12):
                mayus = random.choice(string.ascii_uppercase)
                minus = random.choice(string.ascii_lowercase)
                dig = random.choice(string.digits)
                esp = random.choice('!@#$%^&*()-_=+[]{};:,./?')
                restantes = ''.join(random.choices(string.ascii_letters + string.digits + '!@#$%^&*()-_=+[]{};:,./?', k=longitud-4))
                clave = mayus + minus + dig + esp + restantes
                return ''.join(random.sample(clave, len(clave)))

            temp_password = password if password else generar_clave_temporal()

            # Crear usuario
            usuario = Usuario(
                username=usuario_username,
                email=email,
                correo=email,
                nombre=nombre,
                telefono=telefono,
                id_rol=rol,
                is_active=activo
            )
            usuario.set_password(temp_password)
            usuario.debe_cambiar_clave = True
            usuario.save()
            # Auditoría
            Auditoria.objects.create(
                usuario=user,
                accion='CREAR',
                entidad='Usuario',
                detalle=f'ID: {usuario.id_usuario}, Nombre: {usuario.nombre}'
            )
            action = 'creado'
            # Enviar correo con clave temporal
            try:
                login_url = request.build_absolute_uri(reverse('dashboard:login'))
                html_message = (
                    f"""
                    <div style='font-family:Segoe UI, Arial, sans-serif; color:#111; line-height:1.5;'>
                      <h2 style='margin:0 0 12px; color:#dc2626;'>Dulcería Lilis</h2>
                      <p>Hola <strong>{usuario.nombre}</strong>,</p>
                      <p>Se creó tu acceso al sistema. Aquí están tus credenciales temporales:</p>
                      <div style='background:#f9fafb; border:1px solid #e5e7eb; border-radius:8px; padding:12px; margin:14px 0;'>
                        <p style='margin:0;'><strong>Usuario:</strong> {usuario.username}</p>
                        <p style='margin:6px 0 0;'><strong>Clave temporal:</strong> <code style='font-size:14px;'>{temp_password}</code></p>
                      </div>
                      <p>Accede desde este enlace y cambia tu contraseña en el primer ingreso:</p>
                      <p><a href='{login_url}' style='background:#dc2626; color:#fff; padding:10px 14px; border-radius:6px; text-decoration:none;'>Ir al acceso</a></p>
                      <p style='color:#6b7280; font-size:13px;'>Si no solicitaste esta cuenta, ignora este correo.</p>
                    </div>
                    """
                )
                send_mail(
                    subject='Dulcería Lilis - Tu acceso y clave temporal',
                    message=(
                        f'Hola {usuario.nombre},\n'
                        f'Tu usuario es: {usuario.username}\n'
                        f'Tu clave temporal es: {temp_password}\n\n'
                        f'URL de acceso: {login_url}\n'
                        'Deberás cambiarla en tu primer ingreso.'
                    ),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[usuario.email],
                    fail_silently=True,
                    html_message=html_message,
                )
            except Exception:
                pass
        
        return JsonResponse({
            'success': True,
            'message': f'Usuario {action} correctamente',
            'usuario': {
                'id': usuario.id_usuario,
                'usuario': usuario.username,
                'nombre': usuario.nombre,
                'email': usuario.email
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'errors': {'general': [str(e)]}
        }, status=500)

@login_required
def eliminar_usuario(request, usuario_id):
    """API para eliminar un usuario"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
    
    try:
        from usuarios.models import Usuario
        usuario = get_object_or_404(Usuario, id_usuario=usuario_id)
        
        # No permitir eliminar el usuario actual
        if usuario.id_usuario == user.id_usuario:
            return JsonResponse({
                'success': False,
                'message': 'No puedes eliminar tu propia cuenta'
            })
        
        nombre = usuario.nombre
        usuario.delete()
        # Auditoría
        Auditoria.objects.create(
            usuario=user,
            accion='BORRAR',
            entidad='Usuario',
            detalle=f'ID: {usuario_id}, Nombre: {nombre}'
        )
        return JsonResponse({
            'success': True,
            'message': f'Usuario "{nombre}" eliminado correctamente'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

@login_required
def cambiar_estado_usuario(request, usuario_id):
    """API para activar/desactivar un usuario"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
    
    try:
        from usuarios.models import Usuario
        usuario = get_object_or_404(Usuario, id_usuario=usuario_id)
        
        # No permitir desactivar el usuario actual
        if usuario.id_usuario == user.id_usuario:
            return JsonResponse({
                'success': False,
                'message': 'No puedes cambiar el estado de tu propia cuenta'
            })
        
        # Cambiar estado
        nuevo_estado = request.POST.get('activo', 'false') == 'true'
        usuario.is_active = nuevo_estado
        usuario.save()
        
        estado_texto = 'activado' if nuevo_estado else 'desactivado'
        
        return JsonResponse({
            'success': True,
            'message': f'Usuario "{usuario.nombre}" {estado_texto} correctamente',
            'nuevo_estado': nuevo_estado
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

@login_required
def exportar_usuarios_excel(request):
    """Exportar lista de usuarios a Excel"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    
    try:
        from usuarios.models import Usuario
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        
        # Estilos
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ['ID', 'Usuario', 'Email', 'Nombre', 'Teléfono', 'Rol', 'Estado', 'Último Acceso', 'Fecha Creación']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Obtener usuarios
        usuarios = Usuario.objects.select_related('id_rol').all().order_by('-date_joined')
        
        # Llenar datos
        for row_num, usuario in enumerate(usuarios, 2):
            ws.cell(row=row_num, column=1).value = usuario.id_usuario
            ws.cell(row=row_num, column=2).value = usuario.username
            ws.cell(row=row_num, column=3).value = usuario.email
            ws.cell(row=row_num, column=4).value = usuario.nombre
            ws.cell(row=row_num, column=5).value = usuario.telefono or ''
            ws.cell(row=row_num, column=6).value = usuario.id_rol.nombre if usuario.id_rol else 'Sin rol'
            ws.cell(row=row_num, column=7).value = 'Activo' if usuario.is_active else 'Inactivo'
            ws.cell(row=row_num, column=8).value = usuario.last_login.strftime('%Y-%m-%d %H:%M') if usuario.last_login else 'Nunca'
            ws.cell(row=row_num, column=9).value = usuario.date_joined.strftime('%Y-%m-%d %H:%M') if usuario.date_joined else ''
            
            # Aplicar bordes
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = border
                ws.cell(row=row_num, column=col_num).alignment = Alignment(vertical='center')
        
        # Ajustar ancho de columnas
        column_widths = [8, 20, 30, 25, 15, 20, 12, 20, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename=usuarios_{fecha_actual}.xlsx'
        
        # Guardar y devolver
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        print(f"Error exportando usuarios: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'Error al exportar: {str(e)}'
        }, status=500)

@login_required
def exportar_productos_excel(request):
    """Exportar lista de productos a Excel"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    
    try:
        from productos.models import Producto
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Estilos
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ['ID', 'Nombre', 'Descripción', 'Precio Referencia']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Obtener productos
        productos = Producto.objects.all().order_by('nombre')
        
        # Llenar datos
        for row_num, producto in enumerate(productos, 2):
            ws.cell(row=row_num, column=1).value = producto.id_producto
            ws.cell(row=row_num, column=2).value = producto.nombre
            ws.cell(row=row_num, column=3).value = producto.descripcion
            ws.cell(row=row_num, column=4).value = producto.precio_referencia
            
            # Formatear precio como moneda
            ws.cell(row=row_num, column=4).number_format = '$#,##0'
            
            # Aplicar bordes
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = border
                ws.cell(row=row_num, column=col_num).alignment = Alignment(vertical='center')
        
        # Ajustar ancho de columnas
        column_widths = [8, 35, 50, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename=productos_{fecha_actual}.xlsx'
        
        # Guardar y devolver
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        print(f"Error exportando productos: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'Error al exportar: {str(e)}'
        }, status=500)

@login_required
def obtener_producto(request, producto_id):
    """API para obtener detalles de un producto en formato JSON"""
    import traceback
    
    # Verificar autenticación
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'No autenticado'}, status=401)
    
    try:
        from productos.models import Producto
        from producto_proveedor.models import ProductoProveedor
        
        producto = Producto.objects.get(id_producto=producto_id)
        
        # Obtener proveedores asociados
        proveedores_asociados = ProductoProveedor.objects.filter(
            id_producto=producto
        ).select_related('id_proveedor')
        
        proveedores_data = []
        for pp in proveedores_asociados:
            proveedores_data.append({
                'id': pp.id_proveedor.id_proveedor,
                'nombre': pp.id_proveedor.nombre,
                'contacto': pp.id_proveedor.contacto if hasattr(pp.id_proveedor, 'contacto') else '',
                'precio_acordado': pp.precio_acordado,
                'fecha_registro': pp.fecha_registro.strftime('%Y-%m-%d') if pp.fecha_registro else ''
            })
        
        data = {
            'success': True,
            'producto': {
                'id': producto.id_producto,
                'nombre': producto.nombre,
                'descripcion': producto.descripcion,
                'precio_referencia': producto.precio_referencia,
                'proveedores': proveedores_data
            }
        }
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Producto no encontrado'}, status=404)
    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"Error en obtener_producto: {error_trace}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@login_required
def actualizar_producto(request):
    """API para actualizar un producto"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
    
    # Verificar permisos: Administrador y Bodeguero pueden editar
    user = request.user
    rol_nombre = user.id_rol.nombre if hasattr(user, 'id_rol') and user.id_rol else None
    
    if rol_nombre not in ['Administrador', 'Bodeguero'] and not user.is_superuser:
        return JsonResponse({
            'success': False, 
            'message': 'No tienes permisos para editar productos.'
        }, status=403)
    
    try:
        from productos.models import Producto
        
        producto_id = request.POST.get('product_id')
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        precio_referencia = request.POST.get('precio_referencia')
        
        # Validaciones
        if not producto_id:
            return JsonResponse({'success': False, 'message': 'ID de producto no proporcionado'})
        
        if not nombre:
            return JsonResponse({'success': False, 'message': 'El nombre es requerido'})
        
        if not precio_referencia or int(precio_referencia) <= 0:
            return JsonResponse({'success': False, 'message': 'El precio debe ser mayor a 0'})
        
        # Actualizar producto
        producto = Producto.objects.get(id_producto=producto_id)
        producto.nombre = nombre
        producto.descripcion = descripcion
        producto.precio_referencia = int(precio_referencia)
        producto.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Producto actualizado correctamente',
            'producto': {
                'id': producto.id_producto,
                'nombre': producto.nombre,
                'descripcion': producto.descripcion,
                'precio_referencia': producto.precio_referencia
            }
        })
        
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Producto no encontrado'}, status=404)
    except Exception as e:
        import traceback
        print(f"Error actualizando producto: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)